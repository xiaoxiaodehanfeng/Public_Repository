# -*- coding: utf-8 -*-
# @Time     : 2018/12/17/16:05
# @Author   : Hester Xu
# @Email    : 1603046502@qq.com
# @File     : do_excel.py
# @Software : PyCharm

import openpyxl
import json
from common.request import Request


# excel中每一行数据是一个Case，测试用例封装类
class Case:

    def __init__(self):      # 列写出来，列名作为属性
        self.case_id = None  # 先写属性，后面再赋值
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None

# excel 操作封装类，读取数据
class DoExcel:

    def __init__(self,file_name):
        try:          # 判断文件是否存在
            self.file_name = file_name
            # 打开一个excel文件，返回一个workbook对象实例，把它定义为DoExcel的属性，以便于在这个类的其他地方使用
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            print("{} mot found,please check file path".format(file_name))
            raise e

    def get_cases(self,sheet_name):        # 根据sheet名称，获取在这个sheet里面的所有测试用例数据
        sheet = self.workbook[sheet_name]  # 根据sheet名称获取sheet对象实例
        max_row = sheet.max_row            # 获取sheet最大行数
        cases = []                         # 定义一个空列表，存放即将要放进去的测试用例
        for r in range(2, max_row+1):
            case = Case()         # 每一行是一个Case类的实例，实例里面的属性值就是取到的每一行的单元格的值
            # 列用下标获取，怎么优化？？？
            case.case_id = sheet.cell(row=r,column=1).value    # 取第r行,第1格的值
            case.title = sheet.cell(row=r,column=2).value      # 取第r行,第2格的值
            case.url = sheet.cell(row=r, column=3).value       # 取第r行,第3格的值
            case.data = sheet.cell(row=r, column=4).value      # 取第r行,第4格的值
            case.method = sheet.cell(row=r, column=5).value    # 取第r行,第5格的值
            case.expected = sheet.cell(row=r, column=6).value  # 取第r行,第6格的值
            cases.append(case)                                 # 读完之后，把case的实例放到列表中
        return cases                                          # 循环结束后，返回cases列表

    def get_sheet_names(self):                                # 获取到workbook里面所有的sheet名称的列表
        return self.workbook.sheetnames

    # 根据sheet_name定位到sheet，根据case_id定位到行，取当前行里面的actual单元格，赋值,保存当前工作簿
    def write_back_by_case_id(self,sheet_name,case_id,actual,result):
        sheet = self.workbook[sheet_name]        # 根据sheet名称获取sheet对象实例
        max_row = sheet.max_row                  # 获取sheet最大行数
        for r in range(2,max_row+1):
            case_id_r = sheet.cell(r,1).value    # 获取第r行，第1列，也就是case_id这一列的值
            if case_id_r == case_id:             # 判断excel里面取到的当前行的case_id是否等于传进来的case_id
                sheet.cell(r,7).value = actual   # 写入传进来的actual到当前行的actual列的单元格
                sheet.cell(r,8).value = result   # 写入传进来的result到当前行的result列的单元格
                self.workbook.save(filename=self.file_name)
                break

if __name__ == '__main__':
    # 测试DoExcel类
    do_excel = DoExcel(file_name='../datas/cases_v1.xlsx')       # 实例化一个DoExcel对象
    sheet_names = do_excel.get_sheet_names()                    # 获取到workbook里面所有的sheet名称的列表
    print("sheet名称列表：", sheet_names)
    case_list = ['register']               # 定义一个执行测试用例的列表
    for sheet_name in sheet_names:
        if sheet_name in case_list:   # 当前的sheet_name不在可执行的case_list里面，就不执行
            cases = do_excel.get_cases(sheet_name)   # 测试用例列表,由一个个Case对象/实例组成
            print(sheet_name+"测试用例个数：",len(cases))
            for case in cases:              # 遍历测试用例列表
                print("case信息：",case.__dict__)  # 打印case信息
                data = eval(case.data)    # 从excel中取到的data是一个字符串，把字符串转为字典
                resp = Request(method=case.method,url=case.url,data=data)
                print("status_code：",resp.get_status_code())  # 打印响应码
                resp_dict = resp.get_json()  # 获取请求响应，字典
                print("resp_dict ：",resp_dict)
                resp_text = json.dumps(resp_dict,ensure_ascii=False,indent=4) # 将字典转化为字符串
                print("response：",resp_text)  # 打印响应
                # 判断接口响应和excel里面expected的值是否一致
                if case.expected == resp.get_text():
                    print("result: PASS")
                    # 期望结果与实际结果一致，就写入PASS到result这个单元格
                    do_excel.write_back_by_case_id(sheet_name=sheet_name,case_id=case.case_id,actual=resp.get_text(),result='PASS')
                else:
                    print("result: FAIL")
                    # 期望结果与实际结果不一致，就写入FAIL到result这个单元格
                    do_excel.write_back_by_case_id(sheet_name=sheet_name,case_id=case.case_id,actual=resp.get_text(),result='FAIL')



